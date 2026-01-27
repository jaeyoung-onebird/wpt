"""
엑셀 급여 명세서 생성
"""
import os
from datetime import datetime
from typing import List, Dict
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from utils import calculate_net_pay, format_phone, get_bank_code, extract_yymmdd, now_kst

logger = logging.getLogger(__name__)


class PayrollExporter:
    """급여 명세서 엑셀 생성 클래스"""

    def __init__(self, export_dir: str):
        self.export_dir = export_dir
        os.makedirs(export_dir, exist_ok=True)

    def generate_event_payroll(self, event: Dict, attendances: List[Dict],
                                workers: Dict[int, Dict]) -> str:
        """
        행사별 급여 명세서 생성

        Args:
            event: 행사 정보
            attendances: 출석 기록 리스트
            workers: 근무자 정보 dict (worker_id -> worker)

        Returns:
            str: 생성된 파일 경로
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "급여명세서"

        # 제목
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"[{event['title']}] 급여 명세서"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # 행사 정보
        ws.merge_cells('A2:J2')
        info_cell = ws['A2']
        info_cell.value = f"행사일: {event['event_date']} | 장소: {event['location']}"
        info_cell.alignment = Alignment(horizontal='center')

        # 헤더
        headers = [
            "날짜", "행사명", "이름", "생년월일", "은행", "은행코드",
            "계좌번호", "3.3%공제후금액", "세전금액", "연락처"
        ]

        header_row = 4
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self._get_border()

        # 데이터 입력
        row = header_row + 1
        total_deduction = 0
        total_gross_pay = 0

        # 날짜를 YYMMDD 형식으로 변환
        yymmdd_date = extract_yymmdd(event['event_date'])

        for att in attendances:
            worker = workers.get(att['worker_id'], {})

            # 급여 계산
            gross_pay = event['pay_amount']
            net_pay = int(gross_pay * 0.967)  # 3.3% 공제 후 금액

            total_gross_pay += gross_pay
            total_deduction += net_pay

            # 은행코드 자동 매칭
            bank_name = worker.get('bank_name', '')
            bank_code = get_bank_code(bank_name) if bank_name else ''

            # 데이터
            data_row = [
                yymmdd_date,  # 날짜 (YYMMDD)
                event['title'],  # 행사명
                worker.get('name', ''),  # 이름
                worker.get('birth_date', ''),  # 생년월일
                bank_name,  # 은행
                bank_code,  # 은행코드 (자동 매칭)
                worker.get('bank_account', ''),  # 계좌번호
                net_pay,  # 3.3% 공제 후 금액 (자동 계산)
                gross_pay,  # 세전금액
                format_phone(worker.get('phone', ''))  # 연락처
            ]

            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self._get_border()

                # 금액 컬럼은 숫자 포맷
                if col in [8, 9]:
                    cell.number_format = '#,##0'

            row += 1

        # 합계 행
        total_row = row
        ws.merge_cells(f'A{total_row}:G{total_row}')
        total_cell = ws.cell(row=total_row, column=1)
        total_cell.value = f"합계 ({len(attendances)}명)"
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        total_cell.border = self._get_border()

        for col in range(1, 8):
            ws.cell(row=total_row, column=col).border = self._get_border()
            ws.cell(row=total_row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # 합계 금액 - 3.3% 공제 후 금액
        ws.cell(row=total_row, column=8).value = total_deduction
        ws.cell(row=total_row, column=8).number_format = '#,##0'
        ws.cell(row=total_row, column=8).font = Font(bold=True)
        ws.cell(row=total_row, column=8).border = self._get_border()
        ws.cell(row=total_row, column=8).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # 합계 금액 - 세전금액
        ws.cell(row=total_row, column=9).value = total_gross_pay
        ws.cell(row=total_row, column=9).number_format = '#,##0'
        ws.cell(row=total_row, column=9).font = Font(bold=True)
        ws.cell(row=total_row, column=9).border = self._get_border()
        ws.cell(row=total_row, column=9).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # 마지막 컬럼 (연락처) - 빈칸
        ws.cell(row=total_row, column=10).border = self._get_border()
        ws.cell(row=total_row, column=10).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # 컬럼 너비 조정
        column_widths = {
            'A': 12,  # 날짜
            'B': 20,  # 행사명
            'C': 12,  # 이름
            'D': 12,  # 생년월일
            'E': 12,  # 은행
            'F': 12,  # 은행코드
            'G': 18,  # 계좌번호
            'H': 18,  # 3.3%공제후금액
            'I': 15,  # 세전금액
            'J': 15   # 연락처
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 행 높이
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[header_row].height = 25

        # 파일명 생성 (한국 시간)
        filename = f"급여명세_{event['short_code']}_{now_kst().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_dir, filename)

        # 저장
        wb.save(filepath)
        logger.info(f"Payroll exported: {filepath}")

        return filepath

    def _get_border(self) -> Border:
        """테두리 스타일"""
        thin = Side(border_style="thin", color="000000")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def generate_event_report(self, event: Dict, attendances: List[Dict],
                               workers: Dict[int, Dict],
                               billing_items: List[Dict] = None,
                               expense_items: List[Dict] = None,
                               report_info: Dict = None) -> str:
        """
        행사 보고서 생성 (LK PRIVATE 형식) - 개선된 버전
        - 컬럼 너비 확대 (텍스트 짤림 방지)
        - 부가세 10% 자동계산 수식
        - 합계 자동계산 수식
        - 특이사항 빈칸 추가
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "행사보고서"

        # 기본값 설정
        billing_items = billing_items or []
        expense_items = expense_items or []
        report_info = report_info or {}

        # 스타일 정의
        header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        section_font = Font(bold=True, size=11, color="1F4E79")
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        light_blue_fill = PatternFill(start_color="E7F0FD", end_color="E7F0FD", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        highlight_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')

        current_row = 1

        # ==================== 타이틀 ====================
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'].value = "행 사 보 고 서"
        ws[f'A{current_row}'].font = Font(bold=True, size=20, color="1F4E79")
        ws[f'A{current_row}'].alignment = center_align
        ws.row_dimensions[current_row].height = 35
        current_row += 2

        # ==================== 행사 기본 정보 ====================
        info_data = [
            ("행 사 명", event.get('title', ''), "상 호 명", report_info.get('client_name', '')),
            ("행사일시", event.get('event_date', ''), "", ""),
            ("행사장소", event.get('location', ''), "", ""),
            ("담 당 자", report_info.get('manager', ''), "", ""),
        ]

        for label1, value1, label2, value2 in info_data:
            # 왼쪽 라벨
            ws.merge_cells(f'A{current_row}:B{current_row}')
            ws[f'A{current_row}'].value = label1
            ws[f'A{current_row}'].font = Font(bold=True, size=10)
            ws[f'A{current_row}'].fill = light_blue_fill
            ws[f'A{current_row}'].border = self._get_border()
            ws[f'A{current_row}'].alignment = center_align

            # 왼쪽 값
            ws.merge_cells(f'C{current_row}:F{current_row}')
            ws[f'C{current_row}'].value = value1
            ws[f'C{current_row}'].border = self._get_border()
            ws[f'C{current_row}'].alignment = left_align

            # 오른쪽 라벨/값 (상호명만)
            if label2:
                ws.merge_cells(f'G{current_row}:H{current_row}')
                ws[f'G{current_row}'].value = label2
                ws[f'G{current_row}'].font = Font(bold=True, size=10)
                ws[f'G{current_row}'].fill = light_blue_fill
                ws[f'G{current_row}'].border = self._get_border()
                ws[f'G{current_row}'].alignment = center_align

                ws.merge_cells(f'I{current_row}:L{current_row}')
                ws[f'I{current_row}'].value = value2
                ws[f'I{current_row}'].border = self._get_border()
                ws[f'I{current_row}'].alignment = left_align
            else:
                for col in ['G', 'H', 'I', 'J', 'K', 'L']:
                    ws[f'{col}{current_row}'].border = self._get_border()

            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws[f'{col}{current_row}'].border = self._get_border()

            current_row += 1

        current_row += 1

        # ==================== 청구 상세 내역 ====================
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'].value = "청 구 상 세 내 역"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].fill = section_fill
        ws[f'A{current_row}'].border = self._get_border()
        current_row += 1

        # 청구 헤더
        billing_header_row = current_row
        billing_headers = [("A", "NO"), ("B", "항 목"), ("E", "인원"), ("F", "수량"),
                          ("G", "금 액"), ("I", "부가세(10%)"), ("K", "비 고")]

        for col, header in billing_headers:
            ws[f'{col}{current_row}'].value = header
            ws[f'{col}{current_row}'].font = header_font
            ws[f'{col}{current_row}'].fill = header_fill
            ws[f'{col}{current_row}'].alignment = center_align
            ws[f'{col}{current_row}'].border = self._get_border()

        ws.merge_cells(f'B{current_row}:D{current_row}')
        ws.merge_cells(f'G{current_row}:H{current_row}')
        ws.merge_cells(f'I{current_row}:J{current_row}')
        ws.merge_cells(f'K{current_row}:L{current_row}')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{current_row}'].border = self._get_border()
        current_row += 1

        # 청구 데이터 시작 행 저장
        billing_start_row = current_row
        billing_row_count = max(5, len(billing_items))  # 최소 5행

        for i in range(billing_row_count):
            row = current_row + i
            item = billing_items[i] if i < len(billing_items) else {}

            ws[f'A{row}'].value = i + 1 if i < len(billing_items) else ""
            ws[f'A{row}'].alignment = center_align
            ws[f'A{row}'].border = self._get_border()

            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'].value = item.get('name', '')
            ws[f'B{row}'].border = self._get_border()

            ws[f'E{row}'].value = item.get('count', '') if item else ''
            ws[f'E{row}'].alignment = center_align
            ws[f'E{row}'].border = self._get_border()

            ws[f'F{row}'].value = item.get('quantity', '') if item else ''
            ws[f'F{row}'].alignment = center_align
            ws[f'F{row}'].border = self._get_border()

            # 금액
            ws.merge_cells(f'G{row}:H{row}')
            if item.get('amount'):
                ws[f'G{row}'].value = item.get('amount', 0)
            ws[f'G{row}'].number_format = '#,##0'
            ws[f'G{row}'].alignment = right_align
            ws[f'G{row}'].border = self._get_border()

            # 부가세 (10% 자동계산 수식)
            ws.merge_cells(f'I{row}:J{row}')
            ws[f'I{row}'].value = f'=IF(G{row}="","",ROUND(G{row}*0.1,0))'
            ws[f'I{row}'].number_format = '#,##0'
            ws[f'I{row}'].alignment = right_align
            ws[f'I{row}'].border = self._get_border()

            ws.merge_cells(f'K{row}:L{row}')
            ws[f'K{row}'].value = item.get('note', '')
            ws[f'K{row}'].border = self._get_border()

            for col in ['C', 'D', 'H', 'J', 'L']:
                ws[f'{col}{row}'].border = self._get_border()

        current_row += billing_row_count
        billing_end_row = current_row - 1

        # 청구 합계 (수식 사용)
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].value = "합 계 ( 부가세 포함 )"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = self._get_border()

        ws.merge_cells(f'G{current_row}:L{current_row}')
        ws[f'G{current_row}'].value = f'=SUM(G{billing_start_row}:G{billing_end_row})+SUM(I{billing_start_row}:I{billing_end_row})'
        ws[f'G{current_row}'].number_format = '#,##0'
        ws[f'G{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'].alignment = right_align
        ws[f'G{current_row}'].fill = gray_fill
        ws[f'G{current_row}'].border = self._get_border()

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{current_row}'].border = self._get_border()
        billing_total_row = current_row
        current_row += 2

        # ==================== 경비 내역 ====================
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'].value = "경 비 내 역"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].fill = section_fill
        ws[f'A{current_row}'].border = self._get_border()
        current_row += 1

        # 경비 헤더
        expense_headers = [("A", "NO"), ("B", "사용처"), ("D", "구분"), ("E", "항 목"),
                          ("G", "금 액"), ("I", "지급내용"), ("K", "비 고")]

        for col, header in expense_headers:
            ws[f'{col}{current_row}'].value = header
            ws[f'{col}{current_row}'].font = header_font
            ws[f'{col}{current_row}'].fill = header_fill
            ws[f'{col}{current_row}'].alignment = center_align
            ws[f'{col}{current_row}'].border = self._get_border()

        ws.merge_cells(f'B{current_row}:C{current_row}')
        ws.merge_cells(f'E{current_row}:F{current_row}')
        ws.merge_cells(f'G{current_row}:H{current_row}')
        ws.merge_cells(f'I{current_row}:J{current_row}')
        ws.merge_cells(f'K{current_row}:L{current_row}')
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{current_row}'].border = self._get_border()
        current_row += 1

        # 경비 데이터 시작 행
        expense_start_row = current_row
        expense_row_count = max(4, len(expense_items))  # 최소 4행

        for i in range(expense_row_count):
            row = current_row + i
            item = expense_items[i] if i < len(expense_items) else {}

            ws[f'A{row}'].value = i + 1 if i < len(expense_items) else ""
            ws[f'A{row}'].alignment = center_align
            ws[f'A{row}'].border = self._get_border()

            ws.merge_cells(f'B{row}:C{row}')
            ws[f'B{row}'].value = item.get('place', '')
            ws[f'B{row}'].border = self._get_border()

            ws[f'D{row}'].value = item.get('category', '')
            ws[f'D{row}'].alignment = center_align
            ws[f'D{row}'].border = self._get_border()

            ws.merge_cells(f'E{row}:F{row}')
            ws[f'E{row}'].value = item.get('content', '')
            ws[f'E{row}'].border = self._get_border()

            ws.merge_cells(f'G{row}:H{row}')
            if item.get('amount'):
                ws[f'G{row}'].value = item.get('amount', 0)
            ws[f'G{row}'].number_format = '#,##0'
            ws[f'G{row}'].alignment = right_align
            ws[f'G{row}'].border = self._get_border()

            ws.merge_cells(f'I{row}:J{row}')
            ws[f'I{row}'].value = item.get('payment_method', '')
            ws[f'I{row}'].border = self._get_border()

            ws.merge_cells(f'K{row}:L{row}')
            ws[f'K{row}'].value = item.get('note', '')
            ws[f'K{row}'].border = self._get_border()

            for col in ['C', 'F', 'H', 'J', 'L']:
                ws[f'{col}{row}'].border = self._get_border()

        current_row += expense_row_count
        expense_end_row = current_row - 1

        # 경비 합계 (수식)
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'].value = "합 계"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = self._get_border()

        ws.merge_cells(f'G{current_row}:L{current_row}')
        ws[f'G{current_row}'].value = f'=SUM(G{expense_start_row}:G{expense_end_row})'
        ws[f'G{current_row}'].number_format = '#,##0'
        ws[f'G{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'].alignment = right_align
        ws[f'G{current_row}'].fill = gray_fill
        ws[f'G{current_row}'].border = self._get_border()

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws[f'{col}{current_row}'].border = self._get_border()
        expense_total_row = current_row
        current_row += 2

        # ==================== 인건비 내역 ====================
        ws.merge_cells(f'A{current_row}:L{current_row}')
        ws[f'A{current_row}'].value = "인 건 비 내 역"
        ws[f'A{current_row}'].font = section_font
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].fill = section_fill
        ws[f'A{current_row}'].border = self._get_border()
        current_row += 1

        # 인건비 헤더
        labor_headers = ["NO", "날짜", "근무자", "생년월일", "은행", "은행코드", "계좌번호", "세후지급액", "세전지급액", "공제액", "비고"]
        labor_cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']

        for col, header in zip(labor_cols, labor_headers):
            ws[f'{col}{current_row}'].value = header
            ws[f'{col}{current_row}'].font = header_font
            ws[f'{col}{current_row}'].fill = header_fill
            ws[f'{col}{current_row}'].alignment = center_align
            ws[f'{col}{current_row}'].border = self._get_border()
        ws.merge_cells(f'K{current_row}:L{current_row}')
        ws[f'L{current_row}'].border = self._get_border()
        current_row += 1

        # 인건비 데이터
        labor_start_row = current_row
        total_gross = 0
        total_net = 0
        total_deduction = 0
        yymmdd_date = extract_yymmdd(event.get('event_date', ''))

        for idx, att in enumerate(attendances, 1):
            worker = workers.get(att['worker_id'], {})

            gross_pay = event.get('pay_amount', 0)
            net_pay = int(gross_pay * 0.967)  # 3.3% 공제
            deduction = gross_pay - net_pay

            total_gross += gross_pay
            total_net += net_pay
            total_deduction += deduction

            bank_name = worker.get('bank_name', '')
            bank_code = get_bank_code(bank_name) if bank_name else ''

            labor_data = [
                idx, yymmdd_date, worker.get('name', ''), worker.get('birth_date', ''),
                bank_name, bank_code, worker.get('bank_account', ''),
                net_pay, gross_pay, deduction, ''
            ]

            for col_idx, (col, value) in enumerate(zip(labor_cols, labor_data)):
                ws[f'{col}{current_row}'].value = value
                ws[f'{col}{current_row}'].alignment = center_align
                ws[f'{col}{current_row}'].border = self._get_border()

                if col in ['H', 'I', 'J']:  # 금액 컬럼
                    ws[f'{col}{current_row}'].number_format = '#,##0'
                    ws[f'{col}{current_row}'].alignment = right_align

            ws.merge_cells(f'K{current_row}:L{current_row}')
            ws[f'L{current_row}'].border = self._get_border()
            current_row += 1

        labor_end_row = current_row - 1

        # 인건비 합계
        ws.merge_cells(f'A{current_row}:G{current_row}')
        ws[f'A{current_row}'].value = f"합 계 ({len(attendances)}명)"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = center_align
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = self._get_border()

        ws[f'H{current_row}'].value = f'=SUM(H{labor_start_row}:H{labor_end_row})'
        ws[f'H{current_row}'].number_format = '#,##0'
        ws[f'H{current_row}'].font = Font(bold=True)
        ws[f'H{current_row}'].alignment = right_align
        ws[f'H{current_row}'].fill = gray_fill
        ws[f'H{current_row}'].border = self._get_border()

        ws[f'I{current_row}'].value = f'=SUM(I{labor_start_row}:I{labor_end_row})'
        ws[f'I{current_row}'].number_format = '#,##0'
        ws[f'I{current_row}'].font = Font(bold=True)
        ws[f'I{current_row}'].alignment = right_align
        ws[f'I{current_row}'].fill = gray_fill
        ws[f'I{current_row}'].border = self._get_border()

        ws[f'J{current_row}'].value = f'=SUM(J{labor_start_row}:J{labor_end_row})'
        ws[f'J{current_row}'].number_format = '#,##0'
        ws[f'J{current_row}'].font = Font(bold=True)
        ws[f'J{current_row}'].alignment = right_align
        ws[f'J{current_row}'].fill = gray_fill
        ws[f'J{current_row}'].border = self._get_border()

        ws.merge_cells(f'K{current_row}:L{current_row}')
        ws[f'K{current_row}'].fill = gray_fill
        ws[f'K{current_row}'].border = self._get_border()
        ws[f'L{current_row}'].border = self._get_border()

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{current_row}'].border = self._get_border()
            ws[f'{col}{current_row}'].fill = gray_fill
        labor_total_row = current_row
        current_row += 2

        # ==================== 특이사항 및 정산 ====================
        ws[f'A{current_row}'].value = "※ 특이사항 기재요망"
        ws[f'A{current_row}'].font = Font(bold=True, size=11)

        # 정산 금액 (오른쪽) - 특이사항과 같은 행
        settlement_row = current_row
        ws.merge_cells(f'I{settlement_row}:L{settlement_row}')
        ws[f'I{settlement_row}'].value = "★ 정 산 금 액 ★"
        ws[f'I{settlement_row}'].font = Font(bold=True, size=12)
        ws[f'I{settlement_row}'].fill = highlight_fill
        ws[f'I{settlement_row}'].alignment = center_align
        ws[f'I{settlement_row}'].border = self._get_border()
        for col in ['J', 'K', 'L']:
            ws[f'{col}{settlement_row}'].border = self._get_border()
        settlement_row += 1

        # 수입 (청구합계 참조)
        ws.merge_cells(f'I{settlement_row}:J{settlement_row}')
        ws[f'I{settlement_row}'].value = "수입"
        ws[f'I{settlement_row}'].fill = light_blue_fill
        ws[f'I{settlement_row}'].border = self._get_border()
        ws[f'I{settlement_row}'].alignment = center_align
        ws[f'J{settlement_row}'].border = self._get_border()
        ws.merge_cells(f'K{settlement_row}:L{settlement_row}')
        ws[f'K{settlement_row}'].value = f'=G{billing_total_row}'
        ws[f'K{settlement_row}'].number_format = '#,##0'
        ws[f'K{settlement_row}'].alignment = right_align
        ws[f'K{settlement_row}'].border = self._get_border()
        ws[f'L{settlement_row}'].border = self._get_border()
        revenue_row = settlement_row
        settlement_row += 1

        # 지출 (인건비 + 경비)
        ws.merge_cells(f'I{settlement_row}:J{settlement_row}')
        ws[f'I{settlement_row}'].value = "지출"
        ws[f'I{settlement_row}'].fill = light_blue_fill
        ws[f'I{settlement_row}'].border = self._get_border()
        ws[f'I{settlement_row}'].alignment = center_align
        ws[f'J{settlement_row}'].border = self._get_border()
        ws.merge_cells(f'K{settlement_row}:L{settlement_row}')
        ws[f'K{settlement_row}'].value = f'=I{labor_total_row}+G{expense_total_row}'
        ws[f'K{settlement_row}'].number_format = '#,##0'
        ws[f'K{settlement_row}'].alignment = right_align
        ws[f'K{settlement_row}'].border = self._get_border()
        ws[f'L{settlement_row}'].border = self._get_border()
        expenditure_row = settlement_row
        settlement_row += 1

        # 행사 수익금
        ws.merge_cells(f'I{settlement_row}:J{settlement_row}')
        ws[f'I{settlement_row}'].value = "행사 수익금"
        ws[f'I{settlement_row}'].fill = yellow_fill
        ws[f'I{settlement_row}'].border = self._get_border()
        ws[f'I{settlement_row}'].alignment = center_align
        ws[f'J{settlement_row}'].border = self._get_border()
        ws.merge_cells(f'K{settlement_row}:L{settlement_row}')
        ws[f'K{settlement_row}'].value = f'=K{revenue_row}-K{expenditure_row}'
        ws[f'K{settlement_row}'].number_format = '#,##0'
        ws[f'K{settlement_row}'].font = Font(bold=True)
        ws[f'K{settlement_row}'].alignment = right_align
        ws[f'K{settlement_row}'].fill = yellow_fill
        ws[f'K{settlement_row}'].border = self._get_border()
        ws[f'L{settlement_row}'].border = self._get_border()
        settlement_row += 1

        # 순수익금 (세전 청구금합계 - 세전 인건비지급액 - 경비합계)
        ws.merge_cells(f'I{settlement_row}:J{settlement_row}')
        ws[f'I{settlement_row}'].value = "순수익금"
        ws[f'I{settlement_row}'].fill = highlight_fill
        ws[f'I{settlement_row}'].border = self._get_border()
        ws[f'I{settlement_row}'].alignment = center_align
        ws[f'I{settlement_row}'].font = Font(bold=True)
        ws[f'J{settlement_row}'].border = self._get_border()
        ws.merge_cells(f'K{settlement_row}:L{settlement_row}')
        # 순수익금 = 세전 청구금합계(G열) - 세전 인건비(I열) - 경비합계(G열)
        ws[f'K{settlement_row}'].value = f'=G{billing_total_row}-I{labor_total_row}-G{expense_total_row}'
        ws[f'K{settlement_row}'].number_format = '#,##0'
        ws[f'K{settlement_row}'].font = Font(bold=True, color="FF0000")
        ws[f'K{settlement_row}'].alignment = right_align
        ws[f'K{settlement_row}'].fill = highlight_fill
        ws[f'K{settlement_row}'].border = self._get_border()
        ws[f'L{settlement_row}'].border = self._get_border()

        current_row = settlement_row + 2

        # ==================== 컬럼 너비 조정 (확대) ====================
        column_widths = {
            'A': 5,    # NO
            'B': 14,   # 날짜/항목
            'C': 12,   # 근무자/내용
            'D': 14,   # 주민번호/구분
            'E': 10,   # 은행/인원
            'F': 10,   # 은행코드/수량
            'G': 18,   # 계좌번호/금액
            'H': 14,   # 세후지급액
            'I': 14,   # 세전지급액/부가세
            'J': 12,   # 공제액
            'K': 10,   # 비고
            'L': 10    # 비고 확장
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 파일명 생성
        filename = f"행사보고서_{event.get('short_code', 'report')}_{now_kst().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_dir, filename)

        # 저장
        wb.save(filepath)
        logger.info(f"Event report exported: {filepath}")

        return filepath
